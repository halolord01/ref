Python common code

importing pyhton moduals
import pythonFile  --> no nead for .py if python modual in same directory
--> will have to refrence pythonFile.func()
import pythonFile as pyf  --> refrence the modual as pyf.func()
from modual import function, fun2 --> can direct refrence functions, -->  funciton()
from modual import function as fun --> fun()   | direct execution of funciton renaming
from modual import *    --> complete import as direct refrnece, no modual.
if __name__ == '__main__':  --> will only execute if is in the MAIN modual, not imported

Custom moduals:
customDebug --> a set of debugging tools. methods: debug(),



I/O Functions
print([r]"string", "string") optional r to printout raw text (no '\n' chars)

keyboard input
raw_input("display string: ") --> reds 1 line from input, removes new line char
x = input("sdisplay string: ") --> expects and evaulates valid python input 
print(''' multi line print''')
--------------------------------------------------------------------------------
if sys.platform.lower().startswith('win'):
    IS_WINDOWS = True
else:
    IS_WINDOWS = False

--------------------------------------------------------------------------------

Frequently used commands 	(library functions (https://docs.python.org/2/library/functions.html))

all(iterable),any(iterable) --> returns true if all/any are iterable (like lists)
bin(NUM) --> returns a binary string of the integer
hex(NUM) --> returns xec value of an integer

built in
round(num)
abs(absolute)
int(num), str(num), float(num)

math
5//4 = 1 --> drops decimal
% mod, ** power
math.floor(x) --> round up
math.ceil(x)  --> round down

import statistics       (l = list)
var = statistics.mean(l)
.mean(l)
.mode(l)
.stdev(l)
.variance(l)


------------ Tips and tricks of commonalities ---------

make list into chunks
new_lst = [lst[i:i+SUB] for i in range(0,len(lst), SUB)] --> SUB is the length of each sublist

all(arg in list_ for arg in [list, tuple]) --> all must be true for it to pass...

output of function into lists without making intermediate vars:
from itertools import cycle, chain
a,b = [],[]
for l,v in zip(cycle([a,b]), (chain(*f() for i in range(10))))

or making new lists:
a,b = zip(*[f() for i in range(10)])

=====================================================================================
==================================== Threadding =====================================

from threading import Thread, Event

trigger = Event()
trigger.set()
while trigger.isSet()...
trigger.clear(

    )
thread1 = Thread(target = targetFunction, args = [arg1])
thread1.setDaemon(True)
thread1.start()

# make a blocker - waiting for a thread exit
thread1.join()

NOTES
'''
Threads can share variables to each other and the main thread
- multiprocesses can NOT share variable (I think)
You should have a good way to kill threads...

'''

=====================================================================================


=====================================================================================
Strings
ACCEPTS TUPALS TO CHECK MUTUPLE CONDITIONS
.capitalize() --> adam --> Adam
.casefold() --> better than using .lower() for string matching
.center(len [,fillchar]) --> centers the string padded by whitespace
.ljust(width[fillchar]) --> left justifies string, padded with spaces or fill char
.rjust(...)--> right justify with fill chars / blanks spaces
.count(sub[,start[,end]]) -->number of occurances of substring sub
.encode(encoding='utf-8'[,errors='strict']) --> return encoded bytes object
        --> default is 'strict', raise UnicodeError, 'ignore' 'replace' also work
.endswith(suffix[,start[,end]]) --> returns True if string ends with sufffix
.find(sub [,start,end]) --> return lowest index where 'sub' is found (-1 is retuned if not found)
.format(**mapping)
    'string number {0}'.format(1) --> replaces the {0} in a string with characters
        --> {} --> numeric index, name of keywoard argument
.format_map(mapping) --> similar but can be used directialy, see library...
.index(sub) --> like find, but raises ValueError when sub not found
.isalnum() --> returns true if str is all alpha numericial
.isalpha() .isdecimal() .isdigit() .islower() .isnumeric() .isprintable()
.istitle() --> True If Title Case .isupper()
.join(iterable) --> concentrates the strigns
.lower()
.lstrip([chars]) --> left strips whitesace or definig chars (ALL combinaitions, not prefx)
.partition(sep) --> partions string into tuple: (before sep,sep,after sep)
.replace(old,new[,count]) --> replace old with new, count defines how many times, default is all
.rfind(sub[,start[,end]]) --> return highest index in string where sub found (-1 on failure)
.rindex(...) --> like r find, but raises ValueError when sub not found
.rsplit(sep=None,maxsplit=-1) --behaves like split, but if maxslipt is given at most splits are done from RIGHT
.rstrip([chars]) --> strips Default: whitespace or [chars] from right of str, stops at first non strippable char
.split(sep=None,maxslipt=-1) -->return list of words using sep as delimiter, all, or -1 do all, otherwise maxsplit num
.splitlines([keepends]) --> splits by line, will not keep '\n' unless keepends=True, subset of universal new lines
.startswith(prefix[,start,end]) -->true is str starts with prefix, CAN be tuple of possible prefixes
.strip([chars]) --> strips leading and trailing chars removed, 
.swapcase() --> inverses upper and lower case in string NOTE: not always true that s.swapcase().swapcase() ==s
.title() --> Returns Title Case Version Of The String
.upper()
.zfill(width) --> fills with zeros to length  '45'.zfill(5) --> '00045'

--- formatted printing ---
% as formatting character
print('I am go%(ending)s to the %(num)03d types') % {'ending':'ing', 'num':2}
    -- > 'I am going to the 002 types'
    default flags for x and such
    leadings, '0': padded by 0s specigied (03), '-': left justified (overrides 0s), 
    d/i: decimal int, s: string, x: hex lower, X: hex upper f/F: floating point
'This is my {}'.format('string') --> 'this is my string'
'This {1} {0} string'.format('is', 'my') --> 'This my is string'

Strings and formatting
'{:*^60}'.format('Centered text') --> make a line 60 chars long, filled with *, with 'centered text' in the center


--- TEMPLATES --- 
from string import Template
template = '''I am a $string template'''
temp = Template(template)
temp.substitute(<dict with $replacements as keys>)

','.join(<generator>) --> strings are seperated by commas

import os
==========================================================================================
SCRIPTING

#!/usr/bin/python
put at top of file to run as script from cmd
can then run with ./ from command line

pyinstaller [args] <file> to build as exe
-D --> make one folder
-c command line program
--icon=<icon> --> IconEden has good icons

other EXE installers (Py2Exe, Py2App, cx_Freeze)

==========================================================================================





--------------------------------------------------------------------------------
CLASSES
from abc import ABCMeta, abstractmethod

class Base(object): --> replace object with the name of an class to inherit
    var =2 #This is a class atribute
    def __init__(self,parm1,parm2,parm3="none"):   -->Initilizer funciton
        ClassName.__init__(self,parm1,parm2,parm3) --> Initilizes the inheritance parimiters of class ClassName
        self.parm1 = parm1
        self.parm2 = parm2      --> initilizes the class as self calls
        self.parm3 = parm3
    def fun1(self,parmLocal):
        parmLocal *5  --> noraml funciton in the class call, parlLocal refers to local instance
        self.parm1  --> accessor for paramiter that iniitilized the calss
    @staticmethod
    def forAllClass(): --> does not require a self instance to be called
        print('Same for all')
    @classmethod
    def fun2(cls):
        return cls.var == 2
    
    __metaclass__ = ABCMeta  --> defines class as an abstract
    @abstractmethod --> now Name class can no longer be created as an object and is JUST an abstract
        def objectType(): --> abstract method in the class
            "return a string representing the type of Name object"
            pass

class Child(Base):
    def __init__(self,parm1,parm2):
        Base.__init__(self,pass,to,base,paramiters)  --> also initilizes the base class with given parms

        super([Child,self]).__init__(parms to init...)  --> same as avove code, initilizes the supper (base)
                --> the [Child,self] is not nessicary 
                --> if no child specified, it will initilize one super

        self.parm1 = parm1 ---> 

decorators      --> precedes definitions
@staticmethod   --> does not need a self call
@classmethod    --> a class is passes into the method
@abstractmethod --> this method must be defined in the inherited classes

cl = Name(5,6,7)  --> creares cl as a Name object
cl2 = Name(parm2 = 5, parm1 = 6)   --> creates objects with following insitince, parm3 is default none
cl.fun1(5)  --> calls fun1 in the Name class, same parmiters of cl that was iniitilized
Name.fun1(cl,5)  --> same as above code. calls Name class, with self refrence ofject as cl
Name.forAllClass() == "same for all"
__doc__ --> docstring
__name__ --> funciton name
__class__ --> goes to the class, can string together with obj.__class__.__name__

hasattr(obj,'attrib') --> evaluates whether the obj (or class) has the attribute



funcitons
isinstance(ClassName,object) --> tests wherether object is of type ClassName
setattr(CLASS,ARGUMENT,value) --> set class arguments from inside a function, class = self
    -->for key,val in kwargs.items(): setattr(self,key,val)

-  -  -  -  -  -  -  -  -  -  -  -  -  -  -  -  -  -  -  -  -  -  -  -  -  -  -  -  -  -  -  -  -  -  -  
FUNCTIONS
locals() --> returns a dict of all of the local variables
globals() --> dict of global varables
fun.__code__.co_varnames --> returns a TOUPLE of the function arguments
            .co_argcount --> returns number of arguments
fun.__defaults__ --> retunrs a TOUPLE of the default values stored

def fun(*arguments,**keyArgs): --> VAIRABLE INPUTS, HAS to be in the order of * and **
    arguments --> TOUPLE of variable lenth depending on number of inputs, WITHOUT specified values
    keyArgs --> dictionary of the arguments passes that have values assigned to them
        --> direct refrence as a dict, use keyArgs.items() with 2 inputs to iterate through
fun(1,2,3,4,five=5) --> arguments == (1,2,3,4)



-   -   -   -   -   -   -   -   -   -   -   -   -   -   -   -   -   -   
function    
    __doc__ --> documentation string
    __name__   --> function name
    __qualname__ -->   qualified name
    __code__  -->  code object containing compiled function bytecode
    __defaults__ -->   tuple of any default values for positional or keyword parameters
    __kwdefaults__  -->     mapping of any default values for keyword-only parameters
    __globals__ --> global namespace in which this function was defined
    __annotations__ --> mapping of parameters names to annotations; "return" key is reserved for return annotations.
    __dict__ --> dictionary of the argument keys and keywords


code:
co_argcount --> number of arguments (not including * or ** args)
    co_code --> string of raw compiled bytecode
    co_consts -->   tuple of constants used in the bytecode
    co_filename --> name of file in which this code object was created
    co_firstlineno --> number of first line in Python source code
    co_flags  -->  bitmap: 1=optimized | 2=newlocals | 4=*arg | 8=**arg
    co_lnotab   --> encoded mapping of line numbers to bytecode indices
    co_name --> name with which this code object was defined
    co_names -->   tuple of names of local variables
    co_nlocals --> number of local variables
    co_stacksize  -->  virtual machine stack space required
    co_varnames --> tuple of names of arguments and local variables
--------------------------------------------------------------------------------
Data Structures						[optional]

Lists
len(l) --> returns length of a list
l.append(x) --> add item to end of a list   a[len(a):] = [x]
l.appendleft(x) --> add item to the begenning of a list
l.extend(L) --> extend list by appending all items in given list L
l.extendleft(iterable) --> extend left
l.insert(i,x) --> insert x at index i
l.remove(x) --> remove first occurance of x in l
l.remove(l[i]) --> remove the element in the i index
l.clear () --> removes all elements from the list
l.split() --> splits into a list of words
l.pop([i]) --> removes and returns value at index, returns, if empty, processes the last value
l.popleft() --> pops the first item in a list
l.index(x) --> returns index of first x
l.count(x) --> number times x appears in list
l.sort(cmp=none,key=none,reverse=false)  --> () lowest -> highest; 
l.reverse() --> reverses the elements in the list
l.split() --> returns a list split by the defining char
l.lstrip([]) --> strips the chars off the left side of the str
l[x:y] --> slice from element x to y
l[-1] --> last element in a list 
5 in l --> if 5 is in l returns true
filter(function,iterable) --> returns a filterd list for values of iterable for wich fun is true
range(2,25) -> index 2 to inxex 25
map(fun,ierable) --> maps the fun over the iterable map(str.lower,l)
max(l), min(l)
arrays act like vector addition and subtraction
l1 = l2 --> creates pointer object l2 pointing to l1, SAME OBJECT,   use   l2 = l1[:] for making a new list

[1,2,3] < [1,2,3,4,5]  --> True --> is a as subset of b in any order

Ques
deque is used as a fast que service, using regular lists is really slow, as it has to reubild the list
from collections import deque
queue = deque(['name1','name2','name3'])
queue.popleft() --> first in the que now leaves
queue.append('name4') --> adss 1 to the que

var = set('letters') --> creates a data set of unly unique letters
--> can use math operations on them (+ , -) --> 
(a - b) letters in a but not b 
a|b in a or b     |   a ^ b --> in a or b, but not both     |     a & b --> in both a and b

-    -    -    -    -    -    -    -    -    -    -    -    -    -    -    
USEFUL

sum(len(x) for x in l) --> number of dimentions in a nultidiemntional list
arrVar = [x for x in range(len(ARRAY))] --> creates an array the length of ARRAY with 1,2,3,4...
lst = [function(i) for i in LIST] --> creates a list with the funciton or math applied to each element in list

l[:] = [x for x in LIST if x != '\x00'] --> make new list l from old list without '\x00' char
l = [x.lower() for x in l] --> applies the .lower() function to all in the list
l = [i for sublist in LIST for i in sublist] --> flattens LIST to l
-  -  -  -  -  -  -  -  -  -  -  -  -  -  -  -  -  -  -  -  -  -  -  -  -  -  -
Looping through iteratables

for x,y in enumerate(l): print(x,y) --> x ==0, y === l[0] and so on
for x,y in zip(l1,l2):
range(5) --> iterable --> 0,1,2,3,4
dict.iteriteritems()   (for x,y in dct.iteriteritems(): --> key and value corrosponding)



-  -  -  -  -  -  -  -  -  -  -  -  -  -  -  -  -  -  -  -  -  -  -  -  -  -  -
Dictionaries
dct = {'one':=1, 'three-fiddy':3.50,'dict entry 3':6}
dct['one'] == 1
dct['new'] = x  --> creates a new dict entry 'new' with key x
del dct['key']  --> removed key and values in dct
dct.items() --> have to use it this way to loop through items
dct.values() --> returns all values
dct.keys() --> returns all keys
values can be: nums, strings, dicts, arrays, and FUNCTIONS
dct.keys() --> returns a list of the keys
dict([(key,value),(key,value)]) --> creacts a dict from sequence of key-value pairs
dct = dict(zip(a,b)) --> creates a dictionary where list a and b are molded together
dict(key1=val1,key2=val2) --> creats a dict in the same way
for x,y in dct.items(): print x,y --> prints corrosponding key and value

counters
from collections import counter
l.Counter(I) --> creates a dict with sorted and indexed sols
l.elements(x) --> converts the dict into a list
l.most_common([n]) --> neturns n values of most common
l.count(dict) --> counts indexes 

-   -   -   -   -   -   -   -   -   -   -   -   -   -   -   
Sets

a 'list' without repeating insitinces
{1,2,3,2,4,2,1} --> {1,2,3,4}
basic math and can be used,
a - b --> letter in a but not b
a | b --> leters in either a or b
a & b
a ^ b --> in a or b but not both
-  -  -  -  -  -  -  -  -  -  -  -  -  -  -  -  -  -  -  -  -  -  -  -  -  -  -
Tuples
tuples can not be changed after they are created
faster than lists

var = (1,2,5,6)
varp[0:3] --> 0 to third index of a tuple
cmp(t1,t2) --> compares elements of both touples
len(touple) --> lenth of the touple
max(touple) min(touple) --> returns item with the max value, or min value
tuple(seq) --> converts a list into a touple
num in (tuple) --> returns boolean if num is in the tubple

--------------------------------------------------------------------------------
GENERATORS
a generator is an iterable object that can be changed by * comprehention functions

[u for u in x if 'temp' in u] --> basic if then generator
(x+1 for x in range(50)) --> 50 times will add 1 to each num of 50
(1 if sen[i]==True else 0 for i in range(len(sen))) 
(fun if eval else fun2 for i in range(50)) --> a generator that acts with true and false evaluaitons
comprehention functions:
list(gen) --> [gen] also works
dict(den)
sum(gen)
----------------------------------------------
strings

r.strip -->, default is whitespace
l.rstrip([]) --> strip off of right side of str


--------------------------------------------------------------------------------
Working with files
with open(...) as f:
var = open('FileName.txt','x') --> x is the opening paramiter
	r; read only, r+; reading and writing (cursor at begenning), w; writing,
	w+; reading and writing (will make new file), a; appending (cursor at end)
	a+; appending and reading (c at end),
	b; binary add to each one to open in binary edit mode
f = open('filename.txt','r').read() --> reads the file and sets it to var f
f.close() --> closes the open file
f.flush() -> flush internal buffer
f.isatty() --> true if file is connected to a tty(-like) device
next(f) --> returns next line from the file each time it is called
f.read([count]) --> reads entire file: number of bytes to be read from the begenning of file (optional)
f.readline(size) reads one line from the file
f.readlines(sizehint) --> reads until EOF, returns list containing the lines, size is bytes to read
f.write(string) --> writes the sting to the file var
f.truncate(size) -> truncated to at most that size
f.writelines(sequence) --> writes sequence of strings to file, any iterable object (list of strings)
f.closed (true if closed)
f.mode (access mode openend)
f.name --> file name
f.tell() --> tells current position within file (bytes from begenning)
f.seek(offset, from) --> offset tells how many bytes, from is refrence pos (0 beg, 1 is current, 2 end)
f.count(x) --> counts c in f iterable
open(file,'r').close() --> create a new file 'file' and close it


from collections import Counter
L2 = Counter(myList) --> will create list L2 with sorted data of myList

r = [list(line.rstrip(['x'])) for line in f] --> splits doccument by pieces of 'x' leave empty for all chars
for line in fileinput: line = line.lower() --> works with an individual line MAYBE
lines = f.read().split('') --> might not even work ******
for x in xrange(1,10): print(*lines, sep='\n')  --> DOES SOMEHTING
--------
Parsing operations
.split() --> creates a list sperated by optional delimiters, no del seperates by white space
.rstrip() --> creates a list of individual chars is no delimiter given

import csv  --> useful for parsing
--
EG
parsing collums
for line in f: arrVar = line.split()  --> splits up the line into different lists by colum

NEW PARSING WITH NUMPY

x,y = np.loadtxt(FILENAME,unpack=true,elimiter=',') --> unpacks into a np ARRAY
- - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - 
--- JSON ---

import json
---loading a file---
json.loads(f) --> creates a python object from the string file f
json.load(f) --> 
---writing a file---
json.dump(obj,[seperators=(',',':')]) --> obj to dump is required, sperators are not


BETTER METHODS:
====writing to a file====
json.dump(SAVE,f) --> auto writes to the file f, the contents of save
save a class as a dict with ovj.__dict__ for json steraliziation of class data. Retrieve as a dict.


- - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - 
--- CSV ---
import csv
var = csv.reader(f,delimiter=',') --> gets data as multidemntionql array of strings [[],[]]


--- EXCELL FILES ---
from openpyxl import load_workbook
wb = load_workbook('Name.xlsx')
ws = wb.active   --> ws is now the first worksheet in the workbook
ws['D5:D8'] --> returns the cells in the dims
ws['D5'].value --> returns the value in the cell


wb = load_workbook('Name.xlsx')
ws = wb.active
reading from file (quick and dirty)
roePos = [('D' + str(i + 1)) for i in range(860)]
for i in range(860):
    roe.append(ws[roePos[i]].value)

--- MATLAB ---
# save matlab variables
d = {'matlab_name':var}
scipy.io.savemat('file_name', d)


- - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - 
file operations
import os
.ctermid() --> directory from which python is being run
os.rename(current_file_name, new_file_name)
os.remove('filename.exe')
os.mkdir('name') --> make directory in current directory
os.rmdir('dir') --> removes dir
os.chdir('dirName') -->change directort
os.rename(srd,dst) --> src is the source name, and dst is the final name
os.getcwd() --> dispalys current directory
os.path.isfile(fName) --> tests if fName is in location 
os.getlogin() -->returns the username of the login: pwd.getpwuid(os.getuid())[0] --> usually more powerful
os.get_terminal_size() --> returns size of terminal in a touple (columns, lines)
--> shutil.get_terminal_size() --> high level function that should normally be used
os.chroot(path) --> change the root directory of the current process
os.truncate(path,length)
os.urandom(n) -->string of n byes useable for cryptographicial use
    --> from an os-specific randomness source
.name --> name of OS : mac = 'posix' , possibles 'nt' 'ce' 'java'

os.listdir([path]) --> lists the file names in the directory
os.scandir([path]) --> returns a DirEntry generator object
.name --> filename
.path --> full path name
.is_dir([follow_symlinks=False]) --+ True if directory, false if not []--> follow links in the dir or not
.is_file()
.startswith('CHAR') --> can test the starting char of a string
get_tree_size(f.path) --> outputs the number of child folers


os.times()
.user --> user time
.system --> ststem time
.elapsed --> elapsed time since a fixed point in the past

os.system('cmds') --> run shell commands
or better --> from subprocess import call
call('cmds')

os.getlogin() --> returns username of current user (str)

- - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - 
--- ROS Stuff ---
import rosbag
bag = rosbag.Bag(baf_file)
for topic, msg, t in bag.read_messages(topics = ['topic/one'. 'topic/two']
    #stuff
bag.close()

===========================================================================
===============================================================
argument parsing

Standard: 
sys.argv --> tuple of input arguments, accessable from anywhere
python prog filename * --> will find all matches and pass all through as seperate args
python prog "filename *" --> will pass one argument through with a * character

Advaned : arg parser
import argparse

class Args(object):
    def __init__(self):
        self.arg_parser = argparse.ArgumentParser(prog=name, description=dsc, epilog=epil, add_help=False) #all local variable strings

        optional_args = self.arg_parser.agg_argument_group('Optional Arguments', '')
        optional_args.add_argument('-h', '--help', action='help', help=help_text) #help_text local str var
        optional_args.add_argument('-s', '--sample', type=int, help=sample_help) #name of argument is 'sample'
        
        required_args = self.arg_parser.add_argument_group('Required Arguments', '')
        required_args.add_argument('-i', '--input', type=str, nargs='+', required=True, help=input_help)
    def parse(self):
        args = self.arg_parser.parse_args()
        return args
---
.add_argument parts:
nargs='+' --> the next argument is what the argument wants (ie the string following -i)
    - can also be the number of argumnets following it
    acutal-->
    '?' ; consume 1 argument if avalable
        default=  can be used behind it to set default flag 
    '*' ; consume all following arguments (to next decleration)
    '+' ; consume all following arguments, NEEDS AT LEAST 1
    argparse.REMAINDER --> all other input after this arg is passed through, even - elements 

action='' --> defines the action to be taken
    'store' default useage, stores the value in the arg
    action='store_const', const=num --> creates a default value if not specified
    'store_true', 'store_false' --> defaults true and false, no const specification necessary
    'append' --> allows utuple uses of same -i , and creates a list ie '-i 1 -i 2' --> i = [1,2]
    'append_const' --> constant definition for append, need 'const' field
    'count' --> counts the number of args used, eg: -vvv --> 3
choices=<list> --> argument must be in the list/tuple/list/set
required=<Bool> --> set the arg to required or not
help = <str> --> help string will be displayed for each arg when -h is used


can concat togehter args if none or the last require a value
parsing
-xyz (if at least x,y have defaults)

abreviations allowed (arguments = 'name', 'territorial') (acceptable, -na -ter) etc... unless ambiguious


group = argparse.add_mutually_exclusive_group()    
    group.add_argument...
any arguments added, only 1 can be used or error

.set_defaults(**kwargs) --> will become default arguments even if the arguments were not created with .add_argument
.get_default('key') --> return the default of the key

--- useage ---
(above class saved in Args.py modual)
import Args
arg_parser = Args.Args()
inputArgs = arg_parser.parse()
inputArgs.ARGNAME #ie inputArgs.sample



--------------------------------------------------------------------------------


--------------------------------------------------------------------------------
Looping

for x,y in zip(l1,l2): --> loop through two items at once
for k,v in dictionary.iteriteritems(): --> loop through keys and values in a dictionary

xrange(0,10,2) --> range of 0 to 10, by 2s
reversed(xrange(0,10,2)) --> reverse loop through an item
--------------------------------------------------------------------------------
Plotting
import matplotlib.pyplot as plt
----Plots----
.plot(x,y,label='Line 1',color = 'color',linewidth=2.0) --> basic plotting,
.xlabel('this is the x label',fontsize=X,color='red')
.ylabel('this is the y lable')
.text(x,y,TEXT) -->places text in selected cordinates
.legend([loc = 2]) --> includes a legend in the graph if the plots are labled
.show() --> displays the plot
.figure(1) --> displays the following plots on that figure, more figures are seperate graphs
.subplot(211) --> adds mutuple plots per figure (follow with another plot function)
.title('This is my title') --> titles the graph
clf() --> clears current figure
cla() --> clears current aces
plt.anotate('name',xy=(x,y),xytext=(x1,y1),arrowprops=dict(facecolor='k'mshrink=0.05),)
.savefig('FigureName.png', [bbox_inches = 'tight'])
.figure(figsize=(x,y)) --> args|| figsize, dpi,facecolor,edgecolor,linewidth,tight_layout
----Bar Graphs----
.bar(x,y,...) --> same syntax as .plot, but creates a bar graph
.hist(y,bins,histtype = 'bar',rwidth=0.8) --> histogram; bins is an array of the x bins
.scatter(x,y,marker='*',s=INT) --> scatter plot, optional markers and sizes
.stackplot(x,y,z,a,s,d,colors=['colors','sep','by','colors']) --> creates a stackplot of y-d on x

different chars in text
(r'$\sigma_id=15$') --> sigma
'\alpha', can use most greek letters
http://matplotlib.org/users/mathtext.html#mathtext-tutorial



--- maping colors ---
colormap = plt.cm.gist_ncar
plt.gca().set_color_cycle([colormap(i) for i in np.linspace(0, 0.9, num_plots)])
====================================================================================
Numpy
import numpy as np

arrays --> all have to be the same type of data, and the type can not mutate
var = np.array((1,2,3),(5,2,6)) --> NEEDS OVERARCHING () withing () --> ((VARS))
a = np.array(range(24)) --> makes an array 24 long from 1 - 23
var.ndim --> returns the dimention of the array (above would be a 2 dimentional array)
.shape --> returns the dimentions in a touple (above would be (2,3))
a[:,0] --> returns the first collum (1st index of dim 0)
a.dtype --> retuns the type of array
a.flatten() --> flattens an array ravel() --> also flattens lists, but C-style
a[1][5][0] --> accesses the 1, 5, and 0 indexes in each dimention of array a
a.reshape((3,4,2)) --> respares a into a 3 dimention, 4 row, 2 collum array
a = np.concentrate((x,y,z),axis = int) --> attatches x, y, and z tohether axis 1 is regualar append, 2 cuts dim in half
y = x[:,np.newaxis] --> reverses the row and collums [1,5] ==> [5,1]
.ones((2,3),dtype=int) --> creates a 2x3 array of 1s, DEFAULT IS FLOAT WITHOUT dtype
.zeros((2,3),dtype=int)
.ones_like(a) --> creates an array the same sharpe of a with ones (SAME WITH ZEROS)


Vectors
np.dot(x,y) || np.cross(x,y) --> dotproduct of 2 vectors (arrays) (dot mutuplies np array and np matrix)
        CAN BE USED WITH SIMPLE MATRIXES, [1,2,3]
np.arccos(v) .arcsin .arctan --> inverse sin functions for valc angle (radians)
.linalg.solve(x,y) --> applies scaler vectors y to vector x
====================================================================================
urllib
import urllib.request
import urllib.parse
url.request.urlopen('url') --> gets from the url, page sourse

simple search and responce
url = 'url', values = {'s':'basic','submit':'search'} --> dict of search variables to input
data = urllib.parse.urlencode(values) --> encodes the search into a url string
data = data.encode('utf-8') --> encode with correct encoding
request = urllib.request.Request(url,data) --> what we are going to enter into the page to navagate to
responce = urllib.request.urlopen(request) --> navagating ot the page and opening it
responceData = responce.read() --> reading all of the data from the site.  

human header:    headers= {}
headers['User-Agent'] = 
    "Mozilla/5.0 (X11; Linux i686) AppleWebKit/537.17 (KHTML, like Gecko) Chrome/24.0.1312.27 Safari/537.17"
.Request(url,headers=headers)
====================================================================================
REGULAR EXPRESSIONS
import re

re --> import re
re.search(rex, str, [flag]) --> create MatchObject if found
re.findall(rex, str) --> return list of all matches found

flags: --> options
re.IGNORECASE --> ignore case even with [A-Z]
re.DOTALL --> '.' matches all characters including new lines

Match Object: --> evaluates to true
o.group(NUM) --> returns subgroups that were found

'''
Identifiers
\d ay number            \D --> anything but a number
\s --> space            \S --> anything but a space
\w letters or digits        \W --> anything but a character
. --> any character but a new line  --> \. searches for a period
\b --> whitespace around words

modifiers:
{1,3} --> expecting 1-3  {x}--> x ammount   ex: \d{1-3}  --> didgits 1-3 in length
+ --> Match 1 or more
? --> match 0 or 1
* --> match 0 or more
$ --> match the end of a string
^ --> match the begenning of a string
| --> either or  --> \d{1-3} | \w{1-4}
[] --> range or "vairance"  --> [A-Za-z]  | [1-5a-qA-Z] (find ANY of the following)
            --> [A-Z][a-z] --> capitial A-Z followed by lowercase a-z

White Space charactersL
\n --> new line
\s space
\t tab
\e escape
\f form feed
\r carage return
 
IMPORTANT:
. + * ^ ? [] {} | \ --> have to escape them
    escaping lets you find those actual characters instead of using them as modifiers 

() --> include what is in them in the final result -> r'(\s*)' --> find and include all spaces

Can string together regular expressions
r'\s\w{1-5}\s' --> looking for a space, characters 1-5 in lenght, and then another space

(.*?) --> find everything
'''
FORMAT:
identifier --> modifier


flags
re.fun(r'exp',string,FLAGS)
re.I --> ignore case   -->   re.I|re.M
re.M --> multi line

r'' --> defines that a regular expression is being searched for
.findall() --> find everything that meets the delimiters
        re.findall(r'\d{1,3}',string) --> finds all numbers 1-3 didgits in length
.findall(r'<p>(.*?)</p>') --> find everything between paragraph tags --> any characters of 0+ repitions
var = re.compile(r'regualr expression',flags) --> sets a var name to a regualr expression, lets you use it often
====================================================================================
Serial communication
import serial


====================================================================================
Exceptions
try:    
        x = 5 + "str"
except:    --> except Exception as e:  --> print(e)
        "do this if try fails"
except NameError: ...   --> differnt courses of action for different errors, string together
except: pass --> contine without error
except ZeroDivisionError: --> determins the type of error that will occur
finally: --> will execute finally code after all excepts have been handled (executes first)
raise TypeError("string") --> if executed, will call an error in the code
raise RuntimeError('Error message')
====================================================================================
Debugging (http://www.tutorialspoint.com/python/python_exceptions.htm)

====================================================================================
Shell communication

subprocess.Popen(cmd, shell=True).wait()


import sys
sys.stderr.write('This is an error message') --> creats an error message
sys.stderr.flush() --> should follow error messages
sys.stdout.write() --> regular system output text
.argv --> filename and arguments that are passed to it via command line
      --> creates a list of the imported arguments
sys.platform --> platform the program is running
    if sys.platform == 'win32'





